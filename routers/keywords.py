"""
keywords 路由模块
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Body, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse, HTMLResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from typing import Optional, Dict, Any, List
import json
import time
import os
import secrets
import hashlib
import asyncio
import io
import pandas as pd
from pathlib import Path

from shared import *
from shared import (
    db_manager, cookie_manager, logger, ai_reply_engine,
    verify_token, verify_admin_token, require_auth, get_current_user,
    get_current_user_optional, require_admin, log_with_user,
    generate_token, match_reply, KEYWORDS_MAPPING,
    SESSION_TOKENS, TOKEN_EXPIRE_TIME, security, qr_check_locks,
    qr_check_processed, password_login_sessions, password_login_locks,
    cleanup_qr_check_records, DEFAULT_ADMIN_PASSWORD, ADMIN_USERNAME,
    CAPTCHA_ROUTER_AVAILABLE,
    # Geetest
    geetest_status_store, set_geetest_status, get_geetest_status,
    # API
    API_SECRET_KEY, verify_api_key,
    # Helpers
    serve_frontend, _execute_password_login, process_qr_login_cookies,
    _fallback_save_qr_cookie, _handle_feishu_command, load_keywords,
    check_order_data_completeness,
    # Utils
    qr_login_manager, trans_cookies, image_manager, rate_limit,
    get_file_log_collector,
)

router = APIRouter()

@router.get('/api/default-reply/{cid}')
def get_default_reply_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的默认回复设置（兼容路由）"""
    return get_default_reply(cid, current_user)


@router.put('/api/default-reply/{cid}')
def update_default_reply_compat(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的默认回复设置（兼容路由）"""
    return update_default_reply(cid, reply_data, current_user)


@router.delete('/api/default-reply/{cid}')
def delete_default_reply_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除指定账号的默认回复设置（兼容路由）"""
    return delete_default_reply(cid, current_user)


@router.post('/api/default-reply/{cid}/clear-records')
def clear_default_reply_records_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """清空指定账号的默认回复记录（兼容路由）"""
    return clear_default_reply_records(cid, current_user)


# ------------------------- 飞书命令回调（远程控制机器人）-------------------------


@router.get("/keywords/{cid}")
def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 直接从数据库获取所有关键词（避免重复计算）
    item_keywords = db_manager.get_keywords_with_item_id(cid)

    # 转换为统一格式
    all_keywords = []
    for keyword, reply, item_id in item_keywords:
        all_keywords.append({
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id,
            "type": "item" if item_id else "normal"
        })

    return all_keywords


@router.get("/keywords-with-item-id/{cid}")
def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 获取包含类型信息的关键词
    keywords = db_manager.get_keywords_with_type(cid)

    # 转换为前端需要的格式
    result = []
    for keyword_data in keywords:
        result.append({
            "keyword": keyword_data['keyword'],
            "reply": keyword_data['reply'],
            "item_id": keyword_data['item_id'] or "",
            "type": keyword_data['type'],
            "image_url": keyword_data['image_url']
        })

    return result


@router.post("/keywords/{cid}")
def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    kw_list = [(k, v) for k, v in body.keywords.items()]
    log_with_user('info', f"更新Cookie关键字: {cid}, 数量: {len(kw_list)}", current_user)

    cookie_manager.manager.update_keywords(cid, kw_list)
    log_with_user('info', f"Cookie关键字更新成功: {cid}", current_user)
    return {"msg": "updated", "count": len(kw_list)}


@router.post("/keywords-with-item-id/{cid}")
def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    # 验证数据格式
    keywords_to_save = []
    keyword_set = set()  # 用于检查当前提交的关键词中是否有重复

    for kw_data in body.keywords:
        keyword = kw_data.get('keyword', '').strip()
        reply = kw_data.get('reply', '').strip()
        item_id = kw_data.get('item_id', '').strip() or None

        if not keyword:
            raise HTTPException(status_code=400, detail="关键词不能为空")

        # 检查当前提交的关键词中是否有重复
        keyword_key = f"{keyword}|{item_id or ''}"
        if keyword_key in keyword_set:
            item_id_text = f"（商品ID: {item_id}）" if item_id else "（通用关键词）"
            raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' {item_id_text} 在当前提交中重复")
        keyword_set.add(keyword_key)

        keywords_to_save.append((keyword, reply, item_id))

    # 保存关键词（只保存文本关键词，保留图片关键词）
    try:
        success = db_manager.save_text_keywords_only(cid, keywords_to_save)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词失败")
    except Exception as e:
        error_msg = str(e)

        # 检查是否是图片关键词冲突
        if "已存在（图片关键词）" in error_msg:
            # 直接使用数据库管理器提供的友好错误信息
            raise HTTPException(status_code=400, detail=error_msg)
        elif "UNIQUE constraint failed" in error_msg or "唯一约束冲突" in error_msg:
            # 尝试从错误信息中提取具体的冲突关键词
            conflict_keyword = None
            conflict_type = None

            # 检查是否是数据库管理器抛出的详细错误
            if "关键词唯一约束冲突" in error_msg:
                # 解析详细错误信息：关键词唯一约束冲突: Cookie=xxx, 关键词='xxx', 通用关键词/商品ID: xxx
                import re
                keyword_match = re.search(r"关键词='([^']+)'", error_msg)
                if keyword_match:
                    conflict_keyword = keyword_match.group(1)

                if "通用关键词" in error_msg:
                    conflict_type = "通用关键词"
                elif "商品ID:" in error_msg:
                    item_match = re.search(r"商品ID: ([^\s,]+)", error_msg)
                    if item_match:
                        conflict_type = f"商品关键词（商品ID: {item_match.group(1)}）"

            # 构造用户友好的错误信息
            if conflict_keyword and conflict_type:
                detail_msg = f'关键词 "{conflict_keyword}" （{conflict_type}） 已存在，请使用其他关键词或商品ID'
            elif "keywords.cookie_id, keywords.keyword" in error_msg:
                detail_msg = "关键词重复！该关键词已存在（可能是图片关键词或文本关键词），请使用其他关键词"
            else:
                detail_msg = "关键词重复！请使用不同的关键词或商品ID组合"

            raise HTTPException(status_code=400, detail=detail_msg)
        else:
            log_with_user('error', f"保存关键词时发生未知错误: {error_msg}", current_user)
            raise HTTPException(status_code=500, detail="保存关键词失败")

    log_with_user('info', f"更新Cookie关键字(含商品ID): {cid}, 数量: {len(keywords_to_save)}", current_user)
    return {"msg": "updated", "count": len(keywords_to_save)}


@router.get("/keywords-export/{cid}")
def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出指定账号的关键词为Excel文件"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取关键词数据（包含类型信息）
        keywords = db_manager.get_keywords_with_type(cid)

        # 创建DataFrame，只导出文本类型的关键词
        data = []
        for keyword_data in keywords:
            # 只导出文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                data.append({
                    '关键词': keyword_data['keyword'],
                    '商品ID': keyword_data['item_id'] or '',
                    '关键词内容': keyword_data['reply']
                })

        # 如果没有数据，创建空的DataFrame但保留列名（作为模板）
        if not data:
            df = pd.DataFrame(columns=['关键词', '商品ID', '关键词内容'])
        else:
            df = pd.DataFrame(data)

        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='关键词数据', index=False)

            # 如果是空模板，添加一些示例说明
            if data == []:
                worksheet = writer.sheets['关键词数据']
                # 添加示例数据作为注释（从第2行开始）
                worksheet['A2'] = '你好'
                worksheet['B2'] = ''
                worksheet['C2'] = '您好！欢迎咨询，有什么可以帮助您的吗？'

                worksheet['A3'] = '价格'
                worksheet['B3'] = '123456'
                worksheet['C3'] = '这个商品的价格是99元，现在有优惠活动哦！'

                worksheet['A4'] = '发货'
                worksheet['B4'] = ''
                worksheet['C4'] = '我们会在24小时内发货，请耐心等待。'

                # 设置示例行的样式（浅灰色背景）
                from openpyxl.styles import PatternFill
                gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for row in range(2, 5):
                    for col in range(1, 4):
                        worksheet.cell(row=row, column=col).fill = gray_fill

        output.seek(0)

        # 生成文件名（使用URL编码处理中文）
        from urllib.parse import quote
        if not data:
            filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
        else:
            filename = f"keywords_{cid}_{int(time.time())}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))

        # 返回文件
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"导出关键词失败: {e}")
        raise HTTPException(status_code=500, detail="导出关键词失败")


@router.post("/keywords-import/{cid}")
async def import_keywords(cid: str, file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入Excel文件中的关键词到指定账号"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")

    try:
        # 读取Excel文件
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        # 检查必要的列
        required_columns = ['关键词', '商品ID', '关键词内容']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

        # 获取现有的文本类型关键词（用于比较更新/新增）
        existing_keywords = db_manager.get_keywords_with_type(cid)
        existing_dict = {}
        for keyword_data in existing_keywords:
            # 只考虑文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                keyword = keyword_data['keyword']
                reply = keyword_data['reply']
                item_id = keyword_data['item_id']
                key = f"{keyword}|{item_id or ''}"
                existing_dict[key] = (keyword, reply, item_id)

        # 处理导入数据
        import_data = []
        update_count = 0
        add_count = 0

        def clean_cell_value(value):
            """清理单元格值，处理数字转字符串时的 .0 后缀问题"""
            if pd.isna(value):
                return ''
            # 如果是数字类型，先转为整数（如果是整数值）再转字符串
            if isinstance(value, float) and value == int(value):
                return str(int(value)).strip()
            return str(value).strip()

        for index, row in df.iterrows():
            keyword = clean_cell_value(row['关键词'])
            item_id = clean_cell_value(row['商品ID']) or None
            reply = clean_cell_value(row['关键词内容'])

            if not keyword:
                continue  # 跳过没有关键词的行

            # 检查是否重复
            key = f"{keyword}|{item_id or ''}"
            if key in existing_dict:
                # 更新现有关键词
                update_count += 1
            else:
                # 新增关键词
                add_count += 1

            import_data.append((keyword, reply, item_id))

        if not import_data:
            raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")

        # 保存到数据库（只影响文本关键词，保留图片关键词）
        success = db_manager.save_text_keywords_only(cid, import_data)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词到数据库失败")

        log_with_user('info', f"导入关键词成功: {cid}, 新增: {add_count}, 更新: {update_count}", current_user)

        return {
            "msg": "导入成功",
            "total": len(import_data),
            "added": add_count,
            "updated": update_count
        }

    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel文件为空")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Excel文件格式错误")
    except Exception as e:
        logger.error(f"导入关键词失败: {e}")
        raise HTTPException(status_code=500, detail="导入关键词失败")


@router.post("/keywords/{cid}/image")
async def add_image_keyword(
    cid: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """添加图片关键词"""
    logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}")

    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查参数
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="关键词不能为空")

    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="请选择图片文件")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}, filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 先检查关键词是否已存在
        normalized_item_id = item_id if item_id and item_id.strip() else None
        if db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
            # 删除已保存的图片
            image_manager.delete_image(image_url)
            if normalized_item_id:
                raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' 在商品 '{normalized_item_id}' 中已存在")
            else:
                raise HTTPException(status_code=400, detail=f"通用关键词 '{keyword}' 已存在")

        # 保存图片关键词到数据库
        success = db_manager.save_image_keyword(cid, keyword, image_url, item_id or None)
        if not success:
            # 如果数据库保存失败，删除已保存的图片
            logger.error("数据库保存失败，删除已保存的图片")
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=400, detail="图片关键词保存失败，请稍后重试")

        log_with_user('info', f"添加图片关键词成功: {cid}, 关键词: {keyword}", current_user)

        return {
            "msg": "图片关键词添加成功",
            "keyword": keyword,
            "image_url": image_url,
            "item_id": item_id or None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加图片关键词失败: {e}")
        raise HTTPException(status_code=500, detail="添加图片关键词失败")


@router.get("/keywords-with-type/{cid}")
def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含类型信息的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        keywords = db_manager.get_keywords_with_type(cid)
        return keywords
    except Exception as e:
        logger.error(f"获取关键词列表失败: {e}")
        raise HTTPException(status_code=500, detail="获取关键词列表失败")


@router.delete("/keywords/{cid}/{index}")
def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """根据索引删除关键词"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        # 先获取要删除的关键词信息（用于删除图片文件）
        keywords = db_manager.get_keywords_with_type(cid)
        if 0 <= index < len(keywords):
            keyword_data = keywords[index]

            # 删除关键词
            success = db_manager.delete_keyword_by_index(cid, index)
            if not success:
                raise HTTPException(status_code=400, detail="删除关键词失败")

            # 如果是图片关键词，删除对应的图片文件
            if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
                image_manager.delete_image(keyword_data['image_url'])

            log_with_user('info', f"删除关键词成功: {cid}, 索引: {index}, 关键词: {keyword_data.get('keyword')}", current_user)

            return {"msg": "删除成功"}
        else:
            raise HTTPException(status_code=400, detail="关键词索引无效")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除关键词失败: {e}")
        raise HTTPException(status_code=500, detail="删除关键词失败")


@router.get("/debug/keywords-table-info")
def debug_keywords_table_info(current_user: Dict[str, Any] = Depends(get_current_user)):
    """调试：检查keywords表结构"""
    try:
        import sqlite3
        conn = sqlite3.connect(db_manager.db_path)
        cursor = conn.cursor()

        # 获取表结构信息
        cursor.execute("PRAGMA table_info(keywords)")
        columns = cursor.fetchall()

        # 获取数据库版本
        cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
        version_result = cursor.fetchone()
        db_version = version_result[0] if version_result else "未知"

        conn.close()

        return {
            "db_version": db_version,
            "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns]
        }
    except Exception as e:
        logger.error(f"检查表结构失败: {e}")
        raise HTTPException(status_code=500, detail="检查表结构失败")


# 卡券管理API


@router.put("/item-reply/{cookie_id}/{item_id}")
def update_item_reply(
    cookie_id: str,
    item_id: str,
    data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    更新指定账号和商品的回复内容
    """
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager

        # 验证cookie是否属于用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        reply_content = data.get("reply_content", "").strip()
        if not reply_content:
            raise HTTPException(status_code=400, detail="回复内容不能为空")

        db_manager.update_item_reply(cookie_id=cookie_id, item_id=item_id, reply_content=reply_content)

        return {"message": "商品回复更新成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="更新商品回复失败")


@router.delete("/item-reply/{cookie_id}/{item_id}")
def delete_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    删除指定账号cookie_id和商品item_id的商品回复
    """
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        success = db_manager.delete_item_reply(cookie_id, item_id)
        if not success:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return {"message": "商品回复删除成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="删除商品回复失败")


@router.delete("/item-reply/batch")
async def batch_delete_item_reply(
    req: BatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    批量删除商品回复
    """
    user_id = current_user['user_id']
    from db_manager import db_manager

    # 先校验当前用户是否有权限删除每个cookie对应的回复
    user_cookies = db_manager.get_all_cookies(user_id)
    for item in req.items:
        if item.cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail=f"无权限访问Cookie {item.cookie_id}")

    result = db_manager.batch_delete_item_replies([item.dict() for item in req.items])
    return {
        "success_count": result["success_count"],
        "failed_count": result["failed_count"]
    }


@router.get("/item-reply/{cookie_id}/{item_id}")
def get_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取指定账号cookie_id和商品item_id的商品回复内容
    """
    try:
        user_id = current_user['user_id']
        # 校验cookie_id是否属于当前用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        # 获取指定商品回复
        item_replies = db_manager.get_itemReplays_by_cookie(cookie_id)
        # 找对应item_id的回复
        item_reply = next((r for r in item_replies if r['item_id'] == item_id), None)

        if item_reply is None:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return item_reply

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="获取商品回复失败")


# ------------------------- 数据库备份和恢复接口 -------------------------
